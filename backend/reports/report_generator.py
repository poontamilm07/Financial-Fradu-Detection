# ============================================================
# FILE: backend/reports/report_generator.py
# DESCRIPTION: PDF, Excel, CSV report generation
# ============================================================

import io
import csv
from datetime import datetime
from models.db_models import Transaction


# ─────────────────────────────────────────────
# PDF Report
# ─────────────────────────────────────────────
def generate_pdf(transactions, stats):
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    import io

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)

    y = 750

    # Title
    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, y, "Fraud Detection Report")
    y -= 30

    # Stats
    c.setFont("Helvetica", 10)
    c.drawString(50, y, f"Total Transactions: {stats.get('total', 0)}")
    y -= 20
    c.drawString(50, y, f"Fraud Count: {stats.get('fraud_count', 0)}")
    y -= 20
    c.drawString(50, y, f"Fraud Amount: ₹{stats.get('fraud_amount', 0)}")
    y -= 30

    # Transactions
    for t in transactions[:20]:  # limit for safety
       c.drawString(
           50,
           y,
           f"{str(t.transaction_id or '')} | ₹{float(t.amount or 0)} | {str(t.risk_level or '')} | {'FRAUD' if t.is_fraud else 'SAFE'}"
        )
        
       y -= 15

    if y < 50:
         c.showPage()
         y = 750

    c.save()
    buffer.seek(0)

    return buffer

# ─────────────────────────────────────────────
# Excel Report
# ─────────────────────────────────────────────
def generate_excel(transactions: list, stats: dict) -> io.BytesIO:
    """Generate a styled Excel report using openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference

        wb = openpyxl.Workbook()

        # ── Colour palette ────────────────────────
        FILL_HEADER = PatternFill('solid', fgColor='00F0FF')
        FILL_DARK   = PatternFill('solid', fgColor='1A1A2E')
        FILL_ALT    = PatternFill('solid', fgColor='16213E')
        FILL_FRAUD  = PatternFill('solid', fgColor='2D0000')
        FONT_HEADER = Font(bold=True, color='000000', size=11)
        FONT_TITLE  = Font(bold=True, color='00F0FF', size=14)
        FONT_WHITE  = Font(color='FFFFFF', size=10)
        FONT_RED    = Font(color='FF4444', size=10, bold=True)
        FONT_GREEN  = Font(color='00E676', size=10)
        THIN_BORDER = Border(
            left=Side(style='thin', color='333333'),
            right=Side(style='thin', color='333333'),
            top=Side(style='thin', color='333333'),
            bottom=Side(style='thin', color='333333'),
        )
        CENTER = Alignment(horizontal='center', vertical='center')
        LEFT   = Alignment(horizontal='left',   vertical='center')

        def style_header(ws, row, col_count):
            for c in range(1, col_count + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill   = FILL_HEADER
                cell.font   = FONT_HEADER
                cell.alignment = CENTER
                cell.border = THIN_BORDER

        def style_row(ws, row, col_count, is_fraud=False, alt=False):
            for c in range(1, col_count + 1):
                cell = ws.cell(row=row, column=c)
                cell.fill      = FILL_FRAUD if is_fraud else (FILL_ALT if alt else FILL_DARK)
                cell.font      = FONT_RED   if is_fraud else FONT_WHITE
                cell.alignment = CENTER
                cell.border    = THIN_BORDER

        def autofit(ws):
            for col in ws.columns:
                max_len = max(
                    (len(str(c.value)) for c in col if c.value), default=8
                )
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)

        # ── Sheet 1: Summary ──────────────────────
        ws1 = wb.active
        ws1.title = '📊 Summary'
        ws1.sheet_properties.tabColor = '00F0FF'
        ws1['A1'] = '🔐 AI Financial Fraud Detection Report'
        ws1['A1'].font = FONT_TITLE
        ws1.merge_cells('A1:D1')
        ws1['A2'] = f'Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")} UTC'
        ws1['A2'].font = Font(color='888888', size=10)

        summary_headers = ['Metric', 'Value']
        ws1.append([])
        ws1.append(summary_headers)
        style_header(ws1, ws1.max_row, 2)

        summary_rows = [
            ('Total Transactions',   stats.get('total', 0)),
            ('Fraud Detected',        stats.get('fraud_count', 0)),
            ('Fraud Rate (%)',         round(stats.get('fraud_rate', 0), 2)),
            ('Total Amount (₹)',       round(stats.get('total_amount', 0), 2)),
            ('Fraud Amount (₹)',       round(stats.get('fraud_amount', 0), 2)),
            ('Blocked Transactions',   stats.get('blocked', 0)),
            ('Flagged Transactions',   stats.get('flagged', 0)),
            ('Avg Risk Score',         round(stats.get('avg_risk_score', 0), 4)),
        ]
        for i, row in enumerate(summary_rows):
            ws1.append(list(row))
            style_row(ws1, ws1.max_row, 2, alt=(i % 2 == 1))

        autofit(ws1)

        # ── Sheet 2: Transactions ─────────────────
        ws2 = wb.create_sheet('💳 Transactions')
        ws2.sheet_properties.tabColor = '7B2FF7'
        txn_headers = [
            'ID', 'Transaction ID', 'User ID', 'Amount (₹)', 'Merchant',
            'Category', 'City', 'Status', 'Risk Level', 'Risk Score (%)',
            'Is Fraud', 'Payment Method', 'Transaction Date', 'Created At',
        ]
        ws2.append(txn_headers)
        style_header(ws2, 1, len(txn_headers))

        for i, t in enumerate(transactions):
            row_data = [
                t.id,
                t.transaction_id,
                t.user_id,
                float(t.amount),
                t.merchant_name or '',
                t.merchant_category or '',
                t.city or '',
                (t.status or '').upper(),
                (t.risk_level or '').upper(),
                round(float(t.risk_score or 0) * 100, 2),
                'YES' if t.is_fraud else 'NO',
                t.payment_method or '',
                str(t.transaction_date) if t.transaction_date else '',
                str(t.created_at)[:19] if t.created_at else '',
            ]
            ws2.append(row_data)
            style_row(ws2, i + 2, len(txn_headers), is_fraud=t.is_fraud, alt=(i % 2 == 1))
            # Colour fraud cell
            score_cell = ws2.cell(row=i + 2, column=10)
            if t.is_fraud:
                score_cell.font = FONT_RED
            else:
                score_cell.font = FONT_GREEN

        autofit(ws2)

        # ── Sheet 3: Fraud Only ───────────────────
        ws3 = wb.create_sheet('🚨 Fraud Only')
        ws3.sheet_properties.tabColor = 'FF4444'
        ws3.append(txn_headers)
        style_header(ws3, 1, len(txn_headers))

        fraud_txns = [t for t in transactions if t.is_fraud]
        for i, t in enumerate(fraud_txns):
            ws3.append([
                t.id, t.transaction_id, t.user_id,
                float(t.amount), t.merchant_name or '',
                t.merchant_category or '', t.city or '',
                (t.status or '').upper(), (t.risk_level or '').upper(),
                round(float(t.risk_score or 0) * 100, 2),
                'YES', t.payment_method or '',
                str(t.transaction_date) if t.transaction_date else '',
                str(t.created_at)[:19] if t.created_at else '',
            ])
            style_row(ws3, i + 2, len(txn_headers), is_fraud=True)

        autofit(ws3)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    except ImportError:
        raise RuntimeError(
            'openpyxl is not installed. Run: pip install openpyxl'
        )


# ─────────────────────────────────────────────
# CSV Report
# ─────────────────────────────────────────────
def generate_csv(transactions: list) -> io.BytesIO:
    """Generate a UTF-8 CSV report."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    # Metadata header
    writer.writerow(['# AI Financial Fraud Detection Report'])
    writer.writerow([f'# Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")} UTC'])
    writer.writerow([])

    # Column headers
    writer.writerow([
        'ID', 'Transaction ID', 'User ID', 'Amount (INR)', 'Merchant Name',
        'Merchant Category', 'City', 'Country', 'Status', 'Risk Level',
        'Risk Score', 'Is Fraud', 'Fraud Reason', 'Payment Method',
        'Device Type', 'Transaction Hour', 'Transaction Date', 'Created At',
    ])

    for t in transactions:
        writer.writerow([
            t.id,
            t.transaction_id,
            t.user_id,
            float(t.amount),
            t.merchant_name or '',
            t.merchant_category or '',
            t.city or '',
            t.country or '',
            t.status or '',
            t.risk_level or '',
            round(float(t.risk_score or 0), 4),
            'YES' if t.is_fraud else 'NO',
            (t.fraud_reason or '').replace('\n', ' '),
            t.payment_method or '',
            t.device_type or '',
            t.transaction_hour or '',
            str(t.transaction_date) if t.transaction_date else '',
            str(t.created_at)[:19]  if t.created_at else '',
        ])

    # Encode to bytes
    byte_buffer = io.BytesIO(buffer.getvalue().encode('utf-8-sig'))
    byte_buffer.seek(0)
    return byte_buffer


# ─────────────────────────────────────────────
# Compute summary stats helper
# ─────────────────────────────────────────────
def compute_stats(transactions: list, date_from=None, date_to=None) -> dict:
    total        = len(transactions)
    fraud_list   = [t for t in transactions if t.is_fraud]
    fraud_count  = len(fraud_list)
    total_amount = sum(float(t.amount) for t in transactions)
    fraud_amount = sum(float(t.amount) for t in fraud_list)
    blocked      = sum(1 for t in transactions if t.status == 'blocked')
    flagged      = sum(1 for t in transactions if t.status == 'flagged')
    avg_risk     = (
        sum(float(t.risk_score or 0) for t in transactions) / total
        if total else 0
    )

    return {
        'total':          total,
        'fraud_count':    fraud_count,
        'fraud_rate':     round(fraud_count / total * 100, 2) if total else 0,
        'total_amount':   total_amount,
        'fraud_amount':   fraud_amount,
        'blocked':        blocked,
        'flagged':        flagged,
        'avg_risk_score': round(avg_risk, 4),
        'date_from':      str(date_from) if date_from else 'All time',
        'date_to':        str(date_to)   if date_to   else 'Now',
    }