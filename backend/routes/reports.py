from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from datetime import datetime
from extensions import db
from models.db_models import Transaction, Report
from middleware.decorators import analyst_or_admin_required
import io
import csv
import os
from reports.report_generator import generate_pdf, generate_excel, generate_csv


reports_bp = Blueprint('reports', __name__)





def generate_excel_report(transactions, stats: dict) -> io.BytesIO:
    """Generate Excel report using openpyxl"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary['A1'] = 'AI Fraud Detection Report'
    ws_summary['A1'].font = Font(size=16, bold=True, color='00F0FF')
    ws_summary['A2'] = f'Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")} UTC'

    headers = ['Metric', 'Value']
    summary_rows = [
        ['Total Transactions', stats.get('total', 0)],
        ['Fraud Detected', stats.get('fraud_count', 0)],
        ['Total Amount (₹)', float(stats.get('total_amount', 0))],
        ['Fraud Amount (₹)', float(stats.get('fraud_amount', 0))],
        ['Fraud Rate (%)', round(float(stats.get('fraud_rate', 0)), 2)],
    ]

    for col, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=4, column=col, value=h)
        cell.fill = PatternFill(fill_type='solid', fgColor='00F0FF')
        cell.font = Font(bold=True, color='000000')

    for row_idx, row in enumerate(summary_rows, 5):
        for col, val in enumerate(row, 1):
            cell = ws_summary.cell(row=row_idx, column=col, value=val)
            cell.fill = PatternFill(fill_type='solid', fgColor='1A1A2E')
            cell.font = Font(color='FFFFFF')

    # Transactions Sheet
    ws_txn = wb.create_sheet("Transactions")
    txn_headers = [
        'ID', 'Transaction ID', 'User ID', 'Amount', 'Merchant',
        'Category', 'City', 'Status', 'Risk Level', 'Risk Score',
        'Is Fraud', 'Payment Method', 'Date', 'Created At'
    ]
    for col, h in enumerate(txn_headers, 1):
        cell = ws_txn.cell(row=1, column=col, value=h)
        cell.fill = PatternFill(fill_type='solid', fgColor='00F0FF')
        cell.font = Font(bold=True, color='000000')

    for row_idx, t in enumerate(transactions, 2):
        values = [
            t.id, t.transaction_id, t.user_id, float(t.amount),
            t.merchant_name, t.merchant_category, t.city,
            t.status, t.risk_level, float(t.risk_score or 0),
            'YES' if t.is_fraud else 'NO', t.payment_method,
            str(t.transaction_date), str(t.created_at)
        ]
        fill = PatternFill(
            fill_type='solid',
            fgColor='2D0000' if t.is_fraud else ('1A1A2E' if row_idx % 2 == 0 else '16213E')
        )
        for col, val in enumerate(values, 1):
            cell = ws_txn.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.font = Font(color='FF4444' if t.is_fraud else 'FFFFFF')

    # Auto-fit columns
    for ws in [ws_summary, ws_txn]:
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@reports_bp.route('/generate', methods=['POST'])
@analyst_or_admin_required
def generate_report():
    data = request.get_json()
    report_type = data.get('type', 'excel')  # pdf, excel, csv
    date_from = data.get('date_from')
    date_to = data.get('date_to')

    query = Transaction.query
    # ✅ CLEAN & CORRECT VERSION

    if date_from:
        try:
            date_from = datetime.strptime(date_from, "%Y-%m-%d")
            query = query.filter(Transaction.transaction_date >= date_from)
        except:
            date_from = None

    if date_to:
        try:
            date_to = datetime.strptime(date_to, "%Y-%m-%d")
            query = query.filter(Transaction.transaction_date <= date_to)
        except:
            date_to = None
        
    

    transactions = query.order_by(Transaction.created_at.desc()).all()

    total_amount = sum(float(t.amount) for t in transactions)
    fraud_count = sum(1 for t in transactions if t.is_fraud)
    fraud_amount = sum(float(t.amount) for t in transactions if t.is_fraud)

    stats = {
        'total': len(transactions),
        'fraud_count': fraud_count,
        'total_amount': total_amount,
        'fraud_amount': fraud_amount,
        'fraud_rate': (fraud_count / len(transactions) * 100) if transactions else 0
    }

    try:
        if report_type == 'pdf':
            buffer = generate_pdf(transactions, stats)
            return send_file(
                buffer,
                mimetype='application/pdf',
                as_attachment=True,
                download_name=f'fraud_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            )
        elif report_type == 'excel':
            buffer = generate_excel(transactions, stats)
            return send_file(
                buffer,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'fraud_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
        elif report_type == 'csv':
            buffer = generate_csv(transactions)
            return send_file(
                buffer,
                mimetype='text/csv',
                as_attachment=True,
                download_name=f'fraud_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
            )
            

    except Exception as e:
        print("PDF ERROR:", e)   # 👈 ADD THIS LINE
        return jsonify({'success': False, 'error': str(e)}), 500


@reports_bp.route('/stats', methods=['GET'])
@jwt_required()
def report_stats():
    """Summary stats for reports page"""
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')

    query = Transaction.query
    if date_from:
        query = query.filter(Transaction.transaction_date >= date_from)
    if date_to:
        query = query.filter(Transaction.transaction_date <= date_to)

    transactions = query.all()
    fraud = [t for t in transactions if t.is_fraud]

    return jsonify({
        'success': True,
        'stats': {
            'total': len(transactions),
            'fraud_count': len(fraud),
            'total_amount': sum(float(t.amount) for t in transactions),
            'fraud_amount': sum(float(t.amount) for t in fraud),
            'fraud_rate': len(fraud) / len(transactions) * 100 if transactions else 0
        }
    })
@reports_bp.route('/risk-profiles/<int:user_id>', methods=['GET'])
@jwt_required()
def get_risk_profile(user_id):
    # Example dummy data (replace with real logic later)
    return jsonify({
        "success": True,
        "data": {
            "user_id": user_id,
            "risk_score": 25,
            "risk_level": "LOW",
            "last_updated": str(datetime.utcnow())
        }
    })